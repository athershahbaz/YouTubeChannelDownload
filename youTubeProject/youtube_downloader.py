from pytube import YouTube
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook

def download_video(url, output_path=None):
    downloadStatus = ""
    
    try:
        yt = YouTube(url)
        video = yt.streams.get_highest_resolution()
        if output_path:
            video.download(output_path)
            downloadStatus = "OK"
        else:
            video.download()
            downloadStatus = "OK"
        print("Download completed!")
    except Exception as e:
        print("An error occurred:", str(e))
        downloadStatus = "NOK"
    
    return downloadStatus

if __name__ == "__main__":
    #video_url = input("Enter the YouTube video URL: ")
    batch = 5
    proceedPrompt = ""
    
    excelFile = input("Enter complete file path where videos data is stored: ")
    output_directory = input("Enter the output directory (press Enter for current directory): ").strip()
    
    path = excelFile[0:excelFile.rindex('/')]
    file = "U_" + excelFile[excelFile.rindex('/')+1 : len(excelFile)]
    filePath = path + "/" + file
    
    # Check if file is already there with duplicates dropped then no action needed otherwise remove duplicated and write to new file
    if(file in os.listdir(path)):
        print("file with unique recrods already created")
        df = pd.read_excel(filePath)
    else:
        originalDf = pd.read_excel(excelFile)
        df = originalDf.drop_duplicates(subset=['videoId'])
        df.to_excel(filePath)

    #open excel workbook and load it to Data Frame
    
    wb = load_workbook(filePath)
    ws = wb['Sheet1']
    

    for i in range(len(df)):
        downloadStatus = df.loc[i, "Download"]
        
        if (not pd.isnull(downloadStatus)):
            i += 1
            continue
        else:
            video_url = df.loc[i, "videoId"]
            if output_directory:
                downloadStatus = download_video(video_url, output_directory)
            else:
                downloadStatus = download_video(video_url)
            ws['I' + str(i+2)] = downloadStatus
            i += 1
            batch -= 1
            
            try:
                if (batch == 0):
                    proceedPrompt = input("Current Batch downloaded.Enter numbre of videos to download in next batch.Press q to quit")
                    if(proceedPrompt == 'q'):
                        break
                    else:
                        batch = int(proceedPrompt)
            except Exception as e:
                print("You didn't enter any number. Bye bye")
                break
    
    wb.save(filePath)


    
